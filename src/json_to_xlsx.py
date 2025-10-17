# -*- coding: utf-8 -*-
r"""
JSON sohbet veri setini okunaklı bir XLSX'e dönüştürür.
- 'sohbetler' sayfası: sohbet başına 1 satır (+ Tüm sohbet metni)
- 'mesajlar'  sayfası: mesaj başına 1 satır
- 'özet'      sayfası: yanit_durumu / sentiment / tur / intent dağılımları (adet + %)

Kullanım (Windows):
  cd C:\Users\User\Desktop\mila-ai-eval
  python src\json_to_xlsx.py --in data\raw\20-sohbet-trendyol-mila.json --out outputs\trendyol_mila.xlsx

Gereksinimler:
  pip install pandas openpyxl
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import List, Dict, Any

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


# ---------- Yardımcılar ----------

def _autowidth(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame, min_w: int = 10, max_w: int = 80):
    """Excel sütun genişliklerini, içerik uzunluğuna göre makul şekilde ayarlar."""
    ws = writer.sheets[sheet_name]
    for i, col in enumerate(df.columns, start=1):
        series = df[col].astype(str) if not df.empty else pd.Series([], dtype=str)
        max_len = max([len(str(col))] + [len(s) for s in series.tolist()]) if len(series) else len(str(col))
        width = max(min_w, min(int(max_len * 1.1), max_w))
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width


def _to_datetime(s: Any):
    """Tarih stringlerini (gün/ay/yıl olabilecek) datetime'a çevirir; çeviremezse NaT döner."""
    if s is None or (isinstance(s, str) and s.strip() == ""):
        return pd.NaT
    # dayfirst=True ile Türkçe tarih formatlarını da yakalayalım
    return pd.to_datetime(s, dayfirst=True, errors="coerce")


def load_json(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError("Beklenen kök yapı: liste (list of conversations).")
    return data


# ---------- Dönüştürme ----------

def normalize(convs: List[Dict[str, Any]]):
    sohbet_kayitlari: List[Dict[str, Any]] = []
    mesaj_kayitlari: List[Dict[str, Any]] = []

    for c in convs:
        sohbet_id = c.get("sohbet_id")
        tarih_saat = _to_datetime(c.get("tarih_saat"))
        yanit_durumu = c.get("yanit_durumu")
        sentiment = c.get("sentiment")
        tur = c.get("tur")
        intent = c.get("intent")
        intent_detay = c.get("intent_detay")

        ms = c.get("mesajlar", []) or []
        msg_count = len(ms)

        first_msg_ts = _to_datetime(ms[0].get("timestamp")) if msg_count else pd.NaT
        last_msg_ts  = _to_datetime(ms[-1].get("timestamp")) if msg_count else pd.NaT
        first_customer_text = next(
            (m.get("text") for m in ms if (m.get("sender") or "").lower().startswith("müşteri")),
            None
        )

        # ---- Tüm sohbeti tek metne çevir (multi-line) ----
        def _fmt(m):
            sender = (m.get("sender") or "").strip()
            text = (m.get("text") or "").strip()
            if not sender and not text:
                return ""
            return f"{sender}: {text}"

        conv_text_lines = [line for line in (_fmt(m) for m in ms) if line]
        full_conv_text = "\n".join(conv_text_lines)
        # Çok uç durumlarda Excel hücre sınırına yaklaşmamak için kırpma (opsiyonel):
        # full_conv_text = full_conv_text[:30000]

        sohbet_kayitlari.append({
            "sohbet_id": sohbet_id,
            "tarih_saat": tarih_saat,
            "mesaj_sayisi": msg_count,
            "ilk_mesaj_zaman": first_msg_ts,
            "son_mesaj_zaman": last_msg_ts,
            "ilk_musteri_mesaji": first_customer_text,
            "yanit_durumu": yanit_durumu,
            "sentiment": sentiment,
            "tur": tur,
            "intent": intent,
            "intent_detay": intent_detay,
            "tam_sohbet": full_conv_text,  # <-- yeni alan
        })

        for idx, m in enumerate(ms, start=1):
            mesaj_kayitlari.append({
                "sohbet_id": sohbet_id,
                "mesaj_sira": idx,
                "gonderen": m.get("sender"),
                "zaman": _to_datetime(m.get("timestamp")),
                "metin": m.get("text"),
            })

    df_sohbet = pd.DataFrame(sohbet_kayitlari).sort_values(["tarih_saat", "sohbet_id"], ignore_index=True)
    df_mesaj  = pd.DataFrame(mesaj_kayitlari).sort_values(["sohbet_id", "mesaj_sira"], ignore_index=True)

    # ---- Özet pivotları (0-1 arası yüzde; Excel'de % biçimi uygulanacak) ----
    def _pct_table(s: pd.Series) -> pd.DataFrame:
        cnt = s.value_counts(dropna=False)
        frac = (cnt / max(1, cnt.sum()))
        out = pd.DataFrame({"adet": cnt, "yuzde": frac})
        out.index.name = "kategori"
        return out.reset_index()

    pivots: List[pd.DataFrame] = []
    if not df_sohbet.empty:
        pivots.append(_pct_table(df_sohbet["yanit_durumu"]).assign(metric="yanit_durumu"))
        pivots.append(_pct_table(df_sohbet["sentiment"]).assign(metric="sentiment"))
        pivots.append(_pct_table(df_sohbet["tur"]).assign(metric="tur"))
        pivots.append(_pct_table(df_sohbet["intent"]).assign(metric="intent"))

    df_ozet = pd.concat(pivots, ignore_index=True) if pivots else pd.DataFrame(columns=["kategori","adet","yuzde","metric"])
    return df_sohbet, df_mesaj, df_ozet


def write_excel(out_path: Path, df_sohbet: pd.DataFrame, df_mesaj: pd.DataFrame, df_ozet: pd.DataFrame):
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl", datetime_format="yyyy-mm-dd HH:MM:SS", date_format="yyyy-mm-dd") as writer:
        sohbet_cols = [
            "sohbet_id","tarih_saat","mesaj_sayisi","ilk_mesaj_zaman","son_mesaj_zaman",
            "ilk_musteri_mesaji","yanit_durumu","sentiment","tur","intent","intent_detay",
            "tam_sohbet",  # tüm sohbet metni (multi-line)
        ]
        mesaj_cols = ["sohbet_id","mesaj_sira","gonderen","zaman","metin"]
        ozet_cols  = ["metric","kategori","adet","yuzde"]

        # Sayfaları yaz
        df_sohbet.reindex(columns=sohbet_cols).to_excel(writer, sheet_name="sohbetler", index=False)
        df_mesaj.reindex(columns=mesaj_cols).to_excel(writer, sheet_name="mesajlar", index=False)
        df_ozet.reindex(columns=ozet_cols).to_excel(writer, sheet_name="özet", index=False)

        # Otomatik genişlik
        _autowidth(writer, "sohbetler", df_sohbet.reindex(columns=sohbet_cols), max_w=120)  # tam_sohbet geniş olabilir
        _autowidth(writer, "mesajlar",  df_mesaj.reindex(columns=mesaj_cols), max_w=100)
        _autowidth(writer, "özet",      df_ozet.reindex(columns=ozet_cols),  max_w=30)

        # Filtre satırı
        for name in ["sohbetler", "mesajlar", "özet"]:
            ws = writer.sheets[name]
            ws.auto_filter.ref = ws.dimensions

        # Özet sayfasında yüzde kolonuna % biçimi
        ws_ozet = writer.sheets["özet"]
        if len(df_ozet) > 0:
            for row in range(2, 2 + len(df_ozet)):
                ws_ozet[f"D{row}"].number_format = "0.00%"

        # 'tam_sohbet' sütununda satırların katlanarak görünmesi (wrap text)
        ws_sohbet = writer.sheets["sohbetler"]
        tam_idx = sohbet_cols.index("tam_sohbet") + 1  # 1-based column index
        col_letter = get_column_letter(tam_idx)
        for row in range(1, ws_sohbet.max_row + 1):
            ws_sohbet[f"{col_letter}{row}"].alignment = Alignment(wrap_text=True)


# ---------- CLI ----------

def main():
    ap = argparse.ArgumentParser(description="JSON sohbetlerini okunaklı XLSX'e dönüştür.")
    ap.add_argument("--in", dest="in_path", required=True, help="Girdi JSON yolu (örn: data\\raw\\20-sohbet-trendyol-mila.json)")
    ap.add_argument("--out", dest="out_path", default=None, help="Çıkış XLSX yolu (örn: outputs\\trendyol_mila.xlsx)")
    args = ap.parse_args()

    in_path = Path(args.in_path)
    if not in_path.exists():
        raise SystemExit(f"Girdi bulunamadı: {in_path}")

    out_path = Path(args.out_path) if args.out_path else Path("../outputs/sohbetler.xlsx")

    convs = load_json(in_path)
    df_sohbet, df_mesaj, df_ozet = normalize(convs)
    write_excel(out_path, df_sohbet, df_mesaj, df_ozet)

    print(f"[OK] Yazıldı: {out_path.resolve()}")
    print(f" - sohbetler: {len(df_sohbet):,} satır")
    print(f" - mesajlar : {len(df_mesaj):,} satır")
    print(f" - özet     : {len(df_ozet):,} satır")


if __name__ == "__main__":
    main()
